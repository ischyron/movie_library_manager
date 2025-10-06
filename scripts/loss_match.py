#!/usr/bin/env python3
from __future__ import annotations

import argparse
import difflib
import re
from pathlib import Path
from typing import Iterable, List, Optional, Tuple, Dict

from openpyxl import load_workbook

import scanner as _scanner


def _norm_title(s: str) -> str:
    s = s or ""
    s = re.sub(r"[._]+", " ", s)
    s = re.sub(r"\((\d{4})\)", "", s)
    s = re.sub(r"[^\w\s]", " ", s)
    s = re.sub(r"\s+", " ", s).strip().lower()
    return s


def read_titles_from_excel(path: Path, title_col: Optional[str] = None, year_col: Optional[str] = None) -> List[Tuple[str, Optional[int]]]:
    wb = load_workbook(filename=str(path), read_only=True, data_only=True)
    ws = wb.active
    # header row
    headers = [str(c.value).strip() if c.value is not None else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]
    header_map = {h.lower(): i for i, h in enumerate(headers)}
    def find_col(name_candidates: Iterable[str]) -> Optional[int]:
        for nm in name_candidates:
            i = header_map.get(nm)
            if i is not None:
                return i
        # fuzzy header contains
        for h, i in header_map.items():
            if any(nm in h for nm in name_candidates):
                return i
        return None

    title_idx = find_col([c.lower() for c in ([title_col] if title_col else [])] or ["title", "movie", "name"])
    year_idx = find_col([c.lower() for c in ([year_col] if year_col else [])] or ["year", "release year", "yr"])

    out: List[Tuple[str, Optional[int]]] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        title_val = str(row[title_idx]).strip() if title_idx is not None and row[title_idx] is not None else ""
        if not title_val:
            continue
        year_val = None
        if year_idx is not None and row[year_idx] is not None:
            try:
                year_val = int(str(row[year_idx]).strip()[:4])
            except Exception:
                year_val = None
        # extract year from title if missing
        if year_val is None:
            m = re.search(r"\b(19|20)\d{2}\b", title_val)
            if m:
                try:
                    year_val = int(m.group(0))
                except Exception:
                    year_val = None
        out.append((title_val, year_val))
    return out


def index_library(root: Path) -> List[Tuple[Path, str, Optional[int]]]:
    items: List[Tuple[Path, str, Optional[int]]] = []
    for p in sorted(root.iterdir(), key=lambda x: x.name.lower()):
        if not p.is_dir():
            continue
        # parse title/year from folder name
        # Use scanner._clean_title_and_year logic on folder name
        title, year = _scanner._clean_title_and_year(p.name)  # type: ignore[attr-defined]
        items.append((p, title, year))
    return items


def best_match(title: str, year: Optional[int], lib: List[Tuple[Path, str, Optional[int]]]) -> Tuple[Optional[Path], float]:
    t_norm = _norm_title(title)
    best: Tuple[Optional[Path], float] = (None, 0.0)
    # Pass 1: exact normalized title match, prefer same year
    for p, t, y in lib:
        if _norm_title(t) == t_norm and (year is None or y == year):
            return p, 1.0
    # Pass 2: similarity, prefer same year
    for p, t, y in lib:
        sim = difflib.SequenceMatcher(None, t_norm, _norm_title(t)).ratio()
        # small boost for same year
        if year is not None and y == year:
            sim += 0.05
        if sim > best[1]:
            best = (p, sim)
    return best


def write_loss_csv(out_path: Path, rows: List[Dict[str, str]]) -> None:
    import csv
    out_path.parent.mkdir(parents=True, exist_ok=True)
    with out_path.open("w", newline="") as f:
        w = csv.DictWriter(f, fieldnames=["title", "year", "best_match", "score", "status"])
        w.writeheader()
        for r in rows:
            w.writerow(r)


def main(argv=None) -> int:
    ap = argparse.ArgumentParser(description="Find likely lost movies by fuzzy-matching Excel titles against a Movies folder.")
    ap.add_argument("--excel", default="data/movies-data-loss.xlsx", type=Path, help="Excel file with movie titles")
    ap.add_argument("--root", default="/Volumes/Extreme SSD/Movies/", type=Path, help="Movies folder to scan")
    ap.add_argument("--out", default="data/data_loss_2017.csv", type=Path, help="Output CSV of unmatched titles")
    ap.add_argument("--title-col", default=None, help="Explicit Excel column name for titles")
    ap.add_argument("--year-col", default=None, help="Explicit Excel column name for year")
    ap.add_argument("--min-score", type=float, default=0.86, help="Minimum similarity score to consider a match")
    args = ap.parse_args(argv)

    xl_titles = read_titles_from_excel(args.excel, title_col=args.title_col, year_col=args.year_col)
    lib = index_library(args.root)

    out_rows: List[Dict[str, str]] = []
    for title, year in xl_titles:
        p, score = best_match(title, year, lib)
        if p is None or score < args.min_score:
            out_rows.append({
                "title": title,
                "year": str(year or ""),
                "best_match": "",
                "score": f"{score:.2f}",
                "status": "lost",
            })
        else:
            out_rows.append({
                "title": title,
                "year": str(year or ""),
                "best_match": str(p),
                "score": f"{score:.2f}",
                "status": "found",
            })

    # Filter to genuinely lost
    lost_only = [r for r in out_rows if r["status"] == "lost"]
    write_loss_csv(args.out, lost_only)
    print(f"Wrote {args.out} ({len(lost_only)} lost of {len(out_rows)} total)")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

